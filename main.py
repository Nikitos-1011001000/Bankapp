#!/usr/bin/env python3
"""Анализатор банковских транзакций."""
import csv
import logging
from typing import Any, Dict, List, Optional
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl

from src.bankapp.processing import sort_by_date, process_bank_search
from src.bankapp.utils import load_transactions
from src.bankapp.widget import get_date, mask_account_card


def load_csv(path: str) -> List[Dict[str, Any]]:
    """Загрузка CSV."""
    data: List[Dict[str, Any]] = []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(dict(row))  # type: ignore
        return data
    except FileNotFoundError:
        logging.error(f"CSV файл не найден: {path}")
        return []
    except Exception as e:
        logging.error(f"Ошибка чтения CSV {path}: {e}")
        return []


def load_xlsx(path: str) -> List[Dict[str, Any]]:
    """Загрузка XLSX."""
    data: List[Dict[str, Any]] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws: Optional[Worksheet] = wb.active

        if ws is None:
            logging.error(f"Нет активного листа в {path}")
            return []

        # Безопасное чтение заголовков
        headers: List[str] = []
        for idx, cell in enumerate(ws[1], 1):
            header_value = cell.value
            headers.append(str(header_value) if header_value is not None else f"col_{idx}")

        # Чтение данных
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):  # Пропуск пустых строк
                continue
            row_dict: Dict[str, Any] = {}
            for i, value in enumerate(row):
                header = headers[i] if i < len(headers) else f"col_{i + 1}"
                row_dict[header] = value
            data.append(row_dict)

        return data

    except FileNotFoundError:
        logging.error(f"XLSX файл не найден: {path}")
        return []
    except Exception as e:
        logging.error(f"Ошибка чтения XLSX {path}: {e}")
        return []


def get_status_menu() -> str:
    """Меню статусов."""
    statuses = ["EXECUTED", "CANCELED", "PENDING"]
    print("Введите статус, по которому необходимо выполнить фильтрацию.")
    print(f"Доступные для фильтровки статусы: {', '.join(statuses)}")

    while True:
        user_input = input().strip().upper()
        if user_input in statuses:
            return user_input
        print(f'Статус операции "{user_input}" недоступен.')


def get_date_sorting(filtered: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Сортировка по дате."""
    sort_choice = input("Отсортировать операции по дате? Да/Нет: ").lower()

    if sort_choice in ['да', 'yes', 'y', 'д']:
        direction = input("Отсортировать по возрастанию или по убыванию? ").lower()
        descending = direction in ['по убыванию', 'убыванию', 'desc', 'd']
        print(f"Операции отсортированы {'по убыванию' if descending else 'по возрастанию'}")
        return sort_by_date(filtered, descending)  # type: ignore
    return filtered


def get_rub_filter(transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Фильтр рублевых транзакций."""
    rub_choice = input("Выводить только рублевые транзакции? Да/Нет: ").lower()

    if rub_choice in ['да', 'yes', 'y', 'д']:
        rub_transactions: List[Dict[str, Any]] = []
        for tx in transactions:
            operation_amount = tx.get('operationAmount', {})
            currency = operation_amount.get('currency', {})
            code = currency.get('code', '')
            if code == 'RUB':
                rub_transactions.append(tx)
        print("Показаны только рублевые транзакции")
        return rub_transactions
    return transactions


def get_search_filter(transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Поиск по описанию."""
    search_choice = input("Отфильтровать список транзакций по определенному слову в описании? Да/Нет: ").lower()

    if search_choice in ['да', 'yes', 'y', 'д']:
        search_term = input("Введите слово для поиска: ")
        result = process_bank_search(transactions, search_term)  # type: ignore
        print(f"Найдено {len(result)} операций с '{search_term}'")
        return result
    return transactions


def print_transactions(transactions: List[Dict[str, Any]]) -> None:
    """Красивый вывод транзакций."""
    if not transactions:
        print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")
        return

    print("Распечатываю итоговый список транзакций...")
    print(f"\nВсего банковских операций в выборке: {len(transactions)}")
    print()

    for tx in transactions[:5]:  # Первые 5
        date = get_date(tx.get('date', ''))  # type: ignore
        desc = tx.get('description', '')
        from_info = tx.get('from', '')
        to_info = tx.get('to', '')
        from_to = f"{mask_account_card(from_info)} -> {mask_account_card(to_info)}"
        amount_info = tx.get('operationAmount', {})
        amount = amount_info.get('amount', '0')
        currency = amount_info.get('currency', {}).get('code', 'RUB')

        print(f"{date} {desc}")
        print(f"{from_to}")
        print(f"Сумма: {amount} {currency}")
        print()


def main() -> None:
    """Главная логика."""
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")

    choice = input().strip()

    # 1️⃣ Загрузка данных
    transactions: List[Dict[str, Any]] = []
    if choice == "1":
        print("Для обработки выбран JSON-файл.")
        transactions = load_transactions()
    elif choice == "2":
        print("Для обработки выбран CSV-файл.")
        transactions = load_csv("data/transactions.csv")
    elif choice == "3":
        print("Для обработки выбран XLSX-файл.")
        transactions = load_xlsx("data/transactions_excel.xlsx")
    else:
        print("Неверный выбор!")
        return

    # 2️⃣ Фильтр по статусу
    status = get_status_menu()
    filtered: List[Dict[str, Any]] = [
        tx for tx in transactions if tx.get('state', '').upper() == status
    ]
    print(f'Операции отфильтрованы по статусу "{status}"')

    # 3️⃣ Пошаговая фильтрация
    result = get_date_sorting(filtered)
    result = get_rub_filter(result)
    result = get_search_filter(result)

    # 4️⃣ Вывод
    print_transactions(result)


if __name__ == "__main__":
    main()
